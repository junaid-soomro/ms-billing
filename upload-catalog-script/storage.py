
from openpyxl import load_workbook
from yattag import Doc, indent
import datetime
import string
import random
import time


def id_generator(size=6, chars=string.ascii_uppercase):
       return ''.join(random.choice(chars) for _ in range(size))


dateTime = datetime.datetime.utcnow().strftime(
    '%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'


ECSOD = "On-Demand"

## Products
Basic = "Basic"
Advance = "Advance"
UTM = "UTM"
Enterprise = "Enterprise"
CentOS = "CentOS"
Fedora = "Fedora"
Ubunut = "Ubuntu"
LoadBalancer = "Load-Balancer"
Windows = "Windows"
ECSOD = "On-Demand"
ECS1YNUP = "1-Year-No-UpFront"
ECS1YPUP = "1-Year-Partial-UpFront"
ECS1YFUP = "1-Year-Full-UpFront"
ECS3YNUP = "3-Years-No-UpFront"
ECS3YPUP = "3-Years-Partial-UpFront"
ECS3YFUP = "3-Years-Full-UpFront"
Linux = "Linux"
AVI = "AVI"
Free = "Free"
Paid = "Paid"


OnDemandUnit = 'UsageInHours'
maximum = 10000
i = 0
j = 0
k = 0

wb = load_workbook("Cloud-SKUs-Presales-v0.2.xlsx")
ws = wb.worksheets[2]


rowSize = len(ws['B'])


planId = [id_generator(8) for z in range(1, rowSize-1)]

# for sheets in wb:
cells = ws['A2':'B'+str(rowSize)]
cell2 = ws['C2':'D'+str(rowSize)]


SKU = []
productNames = []
infrastructurePrices = []
FreeorPaid = []
softwarePrices = []



for c1, c2 in cells:
    # print("{0:8} {1:8}".format(c1.value, c2.value.replace(' ', '-')))
    SKU.append(format(c1.value))
    productNames.append(format(c2.value.replace(' ', '_')))

for c3 in cell2:
    # print("{0:8} {1:8} ".format(round(c3.value, 2), round(c4.value, 2)))
    softwarePrices.append(format(round(c4.value, 2)))

# print(FreeorPaid)

doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<catalog xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="CatalogSchema.xsd">'

doc.asis(xml_header)
doc.asis(xml_schema)

with tag('Catalog'):
    with tag('effectiveDate'):
        text(str(dateTime))
    with tag('catalogName'):
        text('DEFAULT')
    with tag('currencies'):
        with tag('currency'):
            text('SAR')
        with tag('currency'):
            text('USD')
    with tag('units'):

        with tag('unit', name=OnDemandUnit):
            text('')
    with tag('products'):
        for ProductIDNames in productNames:
            with tag("product", name=ProductIDNames, prettyName=SKU[k]):
                with tag("category"):
                    text('STANDALONE')
        k = k+1

    with tag('rules'):
        with tag('changePolicy'):
            with tag('changePolicyCase'):
                with tag('policy'):
                    text('IMMEDIATE')
        with tag('changeAlignment'):
            with tag('changeAlignmentCase'):
                with tag('alignment'):
                    text('START_OF_BUNDLE')
        with tag('cancelPolicy'):
            with tag('cancelPolicyCase'):
                with tag('policy'):
                    text('IMMEDIATE')
        with tag('createAlignment'):
            with tag('createAlignmentCase'):
                with tag('alignment'):
                    text('START_OF_BUNDLE')
        with tag('billingAlignment'):
            with tag('billingAlignmentCase'):
                with tag('alignment'):
                    text('ACCOUNT')
        with tag('priceList'):
            with tag('priceListCase'):
                with tag('toPriceList'):
                    text('DEFAULT')

    with tag('plans'):
        for ecsNames in planId:
            with tag('plan', name=ecsNames):
                with tag('product'):
                    if(len(productNames) != 0):
                        text(productNames[i])
                with tag('recurringBillingMode'):
                    text('IN_ADVANCE')
                with tag('finalPhase', type="EVERGREEN"):
                    with tag('duration'):
                        with tag('unit'):
                            text('UNLIMITED')
                        with tag('number'):
                            text('-1')
                            # on Demand Usage logic
                    if(len(productNames) != 0):
                            with tag('recurring'):
                                with tag('billingPeriod'):
                                            if((ECS1YFUP in productNames[i]) or (ECS3YFUP in productNames[i])):
                                                text('ANNUAL')
                                            else:
                                                text('MONTHLY')
                                with tag('recurringPrice'):
                                    with tag('price'):
                                        with tag('currency'):
                                            text('SAR')
                                        with tag('value'):
                                            text(infrastructurePrices[i])
                                    with tag('price'):
                                        with tag('currency'):
                                            text('USD')
                                        with tag('value'):
                                            text(softwarePrices[i])
                                j = j+1
                                i = i+1

                with tag('plansAllowedInBundle'):
                    text('-1')

    with tag('priceLists'):
        with tag('defaultPriceList', name="DEFAULT", prettyName="DEFAULT"):
            with tag('plans'):
                for ecsNames in planId:
                    with tag('plan'):
                        text(ecsNames)

    result = indent(
        doc.getvalue(),
    )

    # print(result)

    with open("storage.xml", "w") as f:
        f.write(result+'\n</catalog>')
        print('Storage Catalog Created')
